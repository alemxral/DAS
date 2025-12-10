"""
Document Parser Service
Extracts variables and data from Excel files with ##variable## format.
"""
import pandas as pd
import re
from typing import List, Dict, Tuple
from pathlib import Path


class DocumentParser:
    """Parses Excel files to extract variables and data."""
    
    VARIABLE_PATTERN = re.compile(r'##([^#]+)##')
    
    def __init__(self):
        """Initialize DocumentParser."""
        pass
    
    def extract_variables(self, text: str) -> List[str]:
        """
        Extract variable names from text with ##variable## format.
        
        Args:
            text: Text containing variables
            
        Returns:
            List of variable names (without ## markers)
        """
        if not isinstance(text, str):
            return []
        return self.VARIABLE_PATTERN.findall(text)
    
    def get_excel_sheets(self, file_path: str) -> List[str]:
        """
        Get list of sheet names from Excel file.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            List of sheet names
        """
        if not Path(file_path).exists():
            return []
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheets = wb.sheetnames
            wb.close()
            return sheets
        except:
            return []
    
    def detect_data_sheet(self, file_path: str) -> str:
        """
        Auto-detect the sheet containing ##variable## headers.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Sheet name with ##variable## headers, or None if not found
        """
        if not Path(file_path).exists():
            return None
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            # Check each sheet for ##variable## pattern in first row
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                
                if first_row:
                    # Check if any cell contains ##variable## pattern
                    for cell_value in first_row:
                        if cell_value and isinstance(cell_value, str) and '##' in cell_value:
                            wb.close()
                            return sheet_name
            
            wb.close()
            return None
        except:
            return None
    
    def detect_template_sheet(self, file_path: str) -> str:
        """
        Auto-detect the sheet containing ##placeholder## for templates.
        
        Args:
            file_path: Path to Excel template file
            
        Returns:
            Sheet name with ##placeholder##, or None if not found
        """
        if not Path(file_path).exists():
            return None
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            # Check each sheet for ##variable## pattern anywhere
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Check all cells in the sheet (limit to first 100 rows for performance)
                for row in ws.iter_rows(max_row=100, values_only=True):
                    for cell_value in row:
                        if cell_value and isinstance(cell_value, str) and '##' in cell_value:
                            wb.close()
                            return sheet_name
            
            wb.close()
            return None
        except:
            return None
    
    def parse_excel_data(self, file_path: str, sheet_name: str = None) -> Dict:
        """
        Parse Excel file to extract variables and data.
        Variables are expected in the first row with ##variable## format.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Specific sheet name (None for first sheet)
            
        Returns:
            Dictionary with structure:
            {
                'variables': List of variable names,
                'data': List of dictionaries (one per row),
                'raw_headers': Original first row values,
                'sheet_name': Sheet name used
            }
            
        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If file format is invalid
        """
        if not Path(file_path).exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            
            # Load workbook with data_only=False to preserve formulas and formatting
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # Get the specified sheet or auto-detect
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
                ws = wb[sheet_name]
            else:
                # Try to auto-detect sheet with ##variable## headers
                detected_sheet = None
                for sname in wb.sheetnames:
                    test_ws = wb[sname]
                    first_row = next(test_ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                    if first_row:
                        for cell_value in first_row:
                            if cell_value and isinstance(cell_value, str) and '##' in cell_value:
                                detected_sheet = sname
                                break
                    if detected_sheet:
                        break
                
                # Use detected sheet or fall back to active sheet
                if detected_sheet:
                    ws = wb[detected_sheet]
                    sheet_name = detected_sheet
                else:
                    ws = wb.active
                    sheet_name = ws.title
            
            # Get all rows
            rows = list(ws.iter_rows(values_only=False))
            if len(rows) < 2:
                raise ValueError("Excel file must have at least 2 rows (header + data)")
            
            # Extract first row (headers with ##variable##)
            header_row = rows[0]
            raw_headers = []
            variables = []
            
            for cell in header_row:
                header_value = cell.value if cell.value else ""
                raw_headers.append(str(header_value))
                
                if header_value:
                    extracted = self.extract_variables(str(header_value))
                    if extracted:
                        variables.append(extracted[0])
                    else:
                        # If no ##variable## format, use the raw value
                        variables.append(str(header_value))
                else:
                    variables.append(f"column_{len(variables)}")
            
            # Parse data rows (skip first row which is header)
            data_rows = []
            for row in rows[1:]:
                row_data = {}
                has_any_value = False  # Track if row has any non-empty value
                
                for col_idx, cell in enumerate(row):
                    if col_idx >= len(variables):
                        break
                    
                    var_name = variables[col_idx]
                    
                    # Handle empty cells
                    if cell.value is None:
                        row_data[var_name] = ""
                        continue
                    
                    # Format cell value based on its number format
                    formatted_value = self._format_cell_value(cell)
                    row_data[var_name] = formatted_value
                    
                    # Check if this cell has a meaningful value
                    if formatted_value and str(formatted_value).strip():
                        has_any_value = True
                
                # Only add row if it has at least one non-empty value
                if has_any_value:
                    data_rows.append(row_data)
            
            wb.close()
            
            return {
                'variables': variables,
                'data': data_rows,
                'raw_headers': raw_headers,
                'sheet_name': sheet_name,
                'total_rows': len(data_rows)
            }
            
        except Exception as e:
            raise ValueError(f"Error parsing Excel file: {str(e)}")
    
    def _format_cell_value(self, cell):
        """
        Format cell value based on its number format to preserve percentages, currency, dates, etc.
        
        Args:
            cell: openpyxl cell object
            
        Returns:
            Formatted string value preserving the original Excel display format
        """
        from datetime import datetime
        from openpyxl.styles.numbers import BUILTIN_FORMATS
        
        value = cell.value
        number_format = cell.number_format
        
        # If value is None or empty, return empty string
        if value is None or value == '':
            return ''
        
        # Try to use Excel's formatted value directly if available
        # This preserves the exact format as displayed in Excel
        try:
            # For dates, use the cell's number format to determine output format
            if isinstance(value, datetime):
                if number_format and number_format != 'General':
                    # Common date formats mapping
                    if 'dd/mm/yyyy' in number_format.lower() or 'd/m/y' in number_format.lower():
                        return value.strftime('%d/%m/%Y')
                    elif 'mm/dd/yyyy' in number_format.lower() or 'm/d/y' in number_format.lower():
                        return value.strftime('%m/%d/%Y')
                    elif 'yyyy-mm-dd' in number_format.lower():
                        return value.strftime('%Y-%m-%d')
                    elif 'dd-mm-yyyy' in number_format.lower():
                        return value.strftime('%d-%m-%Y')
                    elif 'mmmm' in number_format.lower():
                        return value.strftime('%B %d, %Y')
                    # Check for time component
                    if 'h' in number_format.lower() or 's' in number_format.lower():
                        return value.strftime('%d/%m/%Y %H:%M:%S')
                # Default date format preserving day/month/year order
                return value.strftime('%d/%m/%Y')
            
            # Handle numbers with specific formats
            if isinstance(value, (int, float)) and number_format and number_format != 'General':
                # Handle percentages
                if '%' in number_format:
                    # Count decimal places in format
                    if '.0' in number_format:
                        decimals = number_format.count('0', number_format.index('.'))
                        return f"{value * 100:.{decimals}f}%"
                    return f"{value * 100:.0f}%"
                
                # Handle currency formats
                if any(sym in number_format for sym in ['$', '€', '£', '¥', '₹']):
                    currency_symbol = '$'
                    for sym in ['$', '€', '£', '¥', '₹']:
                        if sym in number_format:
                            currency_symbol = sym
                            break
                    
                    # Check for thousands separator
                    if '#,##0' in number_format or '#.##0' in number_format or '# ##0' in number_format:
                        # Determine thousands separator from format
                        thousands_sep = ','
                        if '#.##0' in number_format or '. ' in number_format:
                            thousands_sep = '.'
                        elif '# ##0' in number_format:
                            thousands_sep = ' '
                        
                        # Determine decimal places
                        if '.00' in number_format or ',00' in number_format:
                            return f"{currency_symbol}{value:,.2f}".replace(',', 'TEMP').replace('.', thousands_sep).replace('TEMP', '.')
                        return f"{currency_symbol}{value:,.0f}".replace(',', thousands_sep)
                    return f"{currency_symbol}{value:.2f}"
                
                # Handle numbers with thousands separator
                if '#,##0' in number_format or '#.##0' in number_format or '# ##0' in number_format:
                    # Determine separators from format
                    thousands_sep = ','
                    decimal_sep = '.'
                    
                    if '#.##0' in number_format:
                        thousands_sep = '.'
                        decimal_sep = ','
                    elif '# ##0' in number_format:
                        thousands_sep = ' '
                        decimal_sep = '.'
                    
                    # Count decimal places
                    decimal_places = 0
                    if '.0' in number_format or ',0' in number_format:
                        # Count zeros after decimal point
                        after_decimal = number_format.split('.')[-1] if '.' in number_format else number_format.split(',')[-1]
                        decimal_places = after_decimal.count('0')
                    
                    # Format the number
                    if decimal_places > 0:
                        # Format with decimals
                        formatted = f"{value:,.{decimal_places}f}"
                        # Replace separators
                        if thousands_sep != ',' or decimal_sep != '.':
                            formatted = formatted.replace(',', 'TEMP_THOUSANDS')
                            formatted = formatted.replace('.', decimal_sep)
                            formatted = formatted.replace('TEMP_THOUSANDS', thousands_sep)
                        return formatted
                    else:
                        # Integer format
                        formatted = f"{int(value):,}"
                        if thousands_sep != ',':
                            formatted = formatted.replace(',', thousands_sep)
                        return formatted
            
            # Handle plain integers
            if isinstance(value, int):
                return str(value)
            
            # Handle floats - preserve original precision
            if isinstance(value, float):
                # Check if it's effectively a whole number
                if value.is_integer():
                    return str(int(value))
                # Otherwise preserve the value as-is
                return str(value)
        
        except Exception as e:
            # If formatting fails, fall back to string representation
            pass
        
        # Default: convert to string
        return str(value)
    
    def parse_multiple_sheets(self, file_path: str) -> Dict[str, Dict]:
        """
        Parse all sheets in an Excel file.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Dictionary mapping sheet names to parsed data
        """
        if not Path(file_path).exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        try:
            excel_file = pd.ExcelFile(file_path)
            result = {}
            
            for sheet_name in excel_file.sheet_names:
                try:
                    result[sheet_name] = self.parse_excel_data(file_path, sheet_name)
                except Exception as e:
                    result[sheet_name] = {'error': str(e)}
            
            return result
            
        except Exception as e:
            raise ValueError(f"Error reading Excel file: {str(e)}")
    
    def validate_data_completeness(self, data: Dict) -> Dict:
        """
        Validate that all variables have values in all data rows.
        
        Args:
            data: Parsed data from parse_excel_data
            
        Returns:
            Dictionary with validation results:
            {
                'is_valid': bool,
                'missing_data': List of issues,
                'empty_cells': int
            }
        """
        missing_data = []
        empty_cells = 0
        
        for row_idx, row in enumerate(data['data'], start=2):  # Start at 2 (row 1 is header)
            for var_name, value in row.items():
                if not value or value.strip() == "":
                    missing_data.append({
                        'row': row_idx,
                        'variable': var_name,
                        'message': f"Empty value for '{var_name}' in row {row_idx}"
                    })
                    empty_cells += 1
        
        return {
            'is_valid': len(missing_data) == 0,
            'missing_data': missing_data,
            'empty_cells': empty_cells,
            'total_cells': len(data['variables']) * len(data['data'])
        }
    
    def get_sample_data(self, data: Dict, num_rows: int = 3) -> List[Dict]:
        """
        Get sample data rows for preview.
        
        Args:
            data: Parsed data from parse_excel_data
            num_rows: Number of rows to return
            
        Returns:
            List of sample data rows
        """
        return data['data'][:num_rows]
