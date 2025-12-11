"""
Job Manager Service
Manages job lifecycle, file operations, and output generation.
"""
import os
import json
import zipfile
import shutil
from pathlib import Path
from typing import List, Dict, Optional
from datetime import datetime

from models.job import Job, JobStatus
from services.file_tracker import FileTracker
from services.document_parser import DocumentParser
from services.template_processor import TemplateProcessor
from services.format_converter import FormatConverter


class JobManager:
    """Manages document generation jobs."""
    
    def __init__(self, jobs_dir: str, storage_dir: str):
        """
        Initialize JobManager.
        
        Args:
            jobs_dir: Directory to store job data
            storage_dir: Directory for file tracking
        """
        self.jobs_dir = Path(jobs_dir)
        self.storage_dir = Path(storage_dir)
        self.jobs_dir.mkdir(parents=True, exist_ok=True)
        self.storage_dir.mkdir(parents=True, exist_ok=True)
        
        self.file_tracker = FileTracker(str(self.storage_dir))
        self.document_parser = DocumentParser()
        self.template_processor = TemplateProcessor()
        self.format_converter = FormatConverter()
        
        self.jobs: Dict[str, Job] = {}
        self._load_all_jobs()
    
    def _load_all_jobs(self):
        """Load all jobs from disk."""
        if not self.jobs_dir.exists():
            return
        
        for job_dir in self.jobs_dir.iterdir():
            if job_dir.is_dir():
                metadata_file = job_dir / "metadata.json"
                if metadata_file.exists():
                    try:
                        with open(metadata_file, 'r', encoding='utf-8') as f:
                            job_data = json.load(f)
                            job = Job.from_dict(job_data)
                            self.jobs[job.id] = job
                    except Exception as e:
                        print(f"Error loading job {job_dir.name}: {str(e)}")
    
    def get_job_dir(self, job_id: str) -> Path:
        """Get the directory for a specific job."""
        return self.jobs_dir / job_id
    
    def save_job_metadata(self, job: Job):
        """Save job metadata to disk."""
        job_dir = self.get_job_dir(job.id)
        job_dir.mkdir(parents=True, exist_ok=True)
        
        metadata_file = job_dir / "metadata.json"
        with open(metadata_file, 'w', encoding='utf-8') as f:
            json.dump(job.to_dict(), f, indent=2, ensure_ascii=False)
    
    def create_job(
        self,
        template_path: Optional[str],
        data_path: str,
        output_formats: List[str],
        excel_print_settings: Optional[Dict] = None,
        output_directory: Optional[str] = None,
        filename_variable: str = '##filename##',
        tabname_variable: str = '##tabname##',
        data_sheet: Optional[str] = None,
        template_sheet: Optional[str] = None,
        templates: Optional[List[Dict]] = None
    ) -> Job:
        """
        Create a new job.
        
        Args:
            template_path: Path to template file
            data_path: Path to data file
            output_formats: List of desired output formats
            excel_print_settings: Optional Excel print settings for PDF conversion
            output_directory: Optional custom output directory
            filename_variable: Variable to use for output filenames (default: ##filename##)
            tabname_variable: Variable to use for Excel workbook tab names (default: ##tabname##)
            
        Returns:
            Created Job instance
        """
        job = Job(
            template_path=template_path,
            data_path=data_path,
            output_formats=output_formats
        )
        
        # Add Excel print settings if provided
        if excel_print_settings:
            job.excel_print_settings = excel_print_settings
        
        # Store custom output directory
        if output_directory:
            job.output_directory = output_directory
        
        # Store filename variable
        job.metadata['filename_variable'] = filename_variable
        
        # Store tabname variable
        job.metadata['tabname_variable'] = tabname_variable
        
        # Store sheet names if provided
        if data_sheet:
            job.metadata['data_sheet'] = data_sheet
        if template_sheet:
            job.metadata['template_sheet'] = template_sheet
        
        # Create job directory
        job_dir = self.get_job_dir(job.id)
        job_dir.mkdir(parents=True, exist_ok=True)
        
        # Track and copy files
        try:
            # Handle multi-template mode
            if templates and len(templates) > 0:
                # Process multiple templates
                processed_templates = []
                for idx, tmpl in enumerate(templates):
                    tmpl_path = tmpl.get('path')
                    if not tmpl_path or not os.path.exists(tmpl_path):
                        continue
                    
                    # Track template file
                    template_info = self.file_tracker.track_file(tmpl_path)
                    
                    # Copy to job directory
                    job_template_path = job_dir / f"template_{idx}{Path(tmpl_path).suffix}"
                    shutil.copy2(template_info['local_path'], job_template_path)
                    
                    # Store template info
                    processed_templates.append({
                        'path': str(job_template_path),
                        'original_path': tmpl_path,
                        'priority': tmpl.get('priority', idx),
                        'sheet': tmpl.get('sheet', None),
                        'file_id': template_info['file_id'],
                        'local_path': template_info['local_path']
                    })
                
                # Sort by priority
                processed_templates.sort(key=lambda x: x['priority'])
                job.metadata['templates'] = processed_templates
                
                # Set first template as main for backward compatibility
                if processed_templates:
                    job.template_path = processed_templates[0]['original_path']
                    job.metadata['job_template_path'] = processed_templates[0]['path']
            else:
                # Legacy single template mode
                if template_path:
                    template_info = self.file_tracker.track_file(template_path)
                    job.template_file_id = template_info['file_id']
                    job.local_template_path = template_info['local_path']
                    
                    # Copy files to job directory for processing
                    job_template_path = job_dir / f"template{Path(template_path).suffix}"
                    shutil.copy2(job.local_template_path, job_template_path)
                    job.metadata['job_template_path'] = str(job_template_path)
            
            # Track data file
            data_info = self.file_tracker.track_file(data_path)
            job.data_file_id = data_info['file_id']
            job.local_data_path = data_info['local_path']
            
            job_data_path = job_dir / f"data{Path(data_path).suffix}"
            shutil.copy2(job.local_data_path, job_data_path)
            job.metadata['job_data_path'] = str(job_data_path)
            
        except Exception as e:
            job.update_status(JobStatus.FAILED, f"Error tracking files: {str(e)}")
        
        # Save job
        self.jobs[job.id] = job
        self.save_job_metadata(job)
        
        return job
    
    def get_job(self, job_id: str) -> Optional[Job]:
        """Get a job by ID."""
        return self.jobs.get(job_id)
    
    def get_all_jobs(self) -> List[Job]:
        """Get all jobs."""
        return list(self.jobs.values())
    
    def get_jobs_by_status(self, status: JobStatus) -> List[Job]:
        """Get all jobs with a specific status."""
        return [job for job in self.jobs.values() if job.status == status]
    
    def delete_job(self, job_id: str, force: bool = False) -> Dict[str, any]:
        """
        Delete a job and its associated files.
        
        Args:
            job_id: Job ID
            force: If True, attempt to delete even if processing (with retries)
            
        Returns:
            Dict with 'success' (bool) and optional 'error' (str)
        """
        if job_id not in self.jobs:
            return {'success': False, 'error': 'Job not found'}
        
        job = self.jobs[job_id]
        
        # Check if job is currently processing
        if job.status == JobStatus.PROCESSING and not force:
            return {'success': False, 'error': 'Cannot delete job while it is processing. Please wait for completion or use force delete.'}
        
        # If processing and force=True, try to cancel first
        if job.status == JobStatus.PROCESSING and force:
            if hasattr(job, '_cancel_event'):
                job._cancel_event.set()
                # Wait up to 5 seconds for cancellation
                if hasattr(job, '_thread') and job._thread:
                    job._thread.join(timeout=5.0)
        
        # Delete job directory with retry logic
        job_dir = self.get_job_dir(job_id)
        if job_dir.exists():
            max_retries = 3
            for attempt in range(max_retries):
                try:
                    shutil.rmtree(job_dir)
                    break
                except PermissionError as e:
                    if attempt < max_retries - 1:
                        import time
                        time.sleep(1)  # Wait 1 second before retry
                    else:
                        return {'success': False, 'error': f'Cannot delete job files: {str(e)}. Files may be in use.'}
                except Exception as e:
                    return {'success': False, 'error': f'Error deleting job: {str(e)}'}
        
        # Remove from memory
        del self.jobs[job_id]
        
        return {'success': True}
    
    def process_job(self, job_id: str) -> Job:
        """
        Process a job to generate documents.
        Delegates to appropriate processor based on job type.
        
        Args:
            job_id: Job ID
            
        Returns:
            Updated Job instance
        """
        job = self.get_job(job_id)
        if not job:
            raise ValueError(f"Job not found: {job_id}")
        
        if job.status != JobStatus.PENDING:
            raise ValueError(f"Job {job_id} cannot be processed (status: {job.status.value})")
        
        job.update_status(JobStatus.PROCESSING)
        self.save_job_metadata(job)
        
        try:
            # Delegate to appropriate processor based on job type
            from models.job import JobType
            
            if job.job_type == JobType.SPLIT:
                return self._process_split_job(job)
            elif job.job_type == JobType.MERGE:
                return self._process_merge_job(job)
            else:  # JobType.TEMPLATE
                return self._process_template_job(job)
        
        except Exception as e:
            error_msg = f"Error processing job: {str(e)}"
            print(error_msg)
            job.update_status(JobStatus.FAILED, error_message=error_msg)
            self.save_job_metadata(job)
            raise
    
    def _process_template_job(self, job: 'Job') -> 'Job':
        """
        Process a template job (original functionality).
        
        Args:
            job: Job instance
            
        Returns:
            Updated Job instance
        """
        import time
        try:
            # Parse data file with optional sheet name
            data_sheet = job.metadata.get('data_sheet', None)
            data_result = self.document_parser.parse_excel_data(
                job.metadata['job_data_path'],
                sheet_name=data_sheet
            )
            job.total_records = data_result['total_rows']
            
            # Create output directory
            output_dir = self.get_job_dir(job.id) / "outputs"
            output_dir.mkdir(parents=True, exist_ok=True)
            
            # Initialize for batched saves
            last_save_time = time.time()
            save_interval = 5  # Save every 5 seconds
            save_count = 10  # Or every 10 rows
            rows_since_save = 0
            
            # Process each data row
            for idx, row_data in enumerate(data_result['data'], start=1):
                # Check for cancellation
                if hasattr(job, '_cancel_event') and job._cancel_event and job._cancel_event.is_set():
                    print(f"Job {job.id}: Cancellation requested, stopping at row {idx}")
                    job.update_status(JobStatus.CANCELLED, error_message="Job cancelled by user")
                    self.save_job_metadata(job)
                    return job
                
                try:
                    # Determine output filename
                    # Check for ##filename## variable or custom filename variable in job metadata
                    filename_var = job.metadata.get('filename_variable', '##filename##')
                    # Remove the ## markers to get the key
                    filename_key = filename_var.replace('##', '')
                    
                    # Try to get custom filename from row data
                    custom_filename = None
                    if filename_key in row_data and row_data[filename_key]:
                        custom_filename = str(row_data[filename_key]).strip()
                        # Sanitize filename - remove invalid characters
                        invalid_chars = '<>:"/\\|?*'
                        for char in invalid_chars:
                            custom_filename = custom_filename.replace(char, '_')
                    
                    # Use custom filename or default to processed_{idx}
                    if custom_filename:
                        base_filename = custom_filename
                    else:
                        base_filename = f"processed_{idx}"
                    
                    # Check if using multi-template mode
                    templates_list = job.metadata.get('templates', [])
                    processed_docs = []
                    
                    if templates_list and len(templates_list) > 0:
                        # Multi-template mode: process all templates for this row
                        print(f"Processing row {idx} with {len(templates_list)} templates...")
                        
                        for tmpl_idx, template_config in enumerate(templates_list):
                            template_path = template_config['path']
                            template_sheet = template_config.get('sheet', None)
                            template_ext = Path(template_path).suffix
                            
                            # Create output with template index to keep them separate
                            processed_doc = output_dir / f"{base_filename}_tmpl{tmpl_idx}{template_ext}"
                            
                            print(f"Row {idx}, Template {tmpl_idx+1}/{len(templates_list)}: Processing {Path(template_path).name}...")
                            
                            self.template_processor.process_template(
                                template_path,
                                row_data,
                                str(processed_doc),
                                sheet_name=template_sheet
                            )
                            
                            if not processed_doc.exists():
                                raise RuntimeError(f"Failed to create document from template {tmpl_idx}: {processed_doc}")
                            
                            processed_docs.append({
                                'path': processed_doc,
                                'priority': template_config['priority'],
                                'template_idx': tmpl_idx
                            })
                        
                        print(f"Row {idx}: All templates processed successfully")
                    else:
                        # Legacy single template mode
                        template_ext = Path(job.metadata['job_template_path']).suffix
                        processed_doc = output_dir / f"{base_filename}{template_ext}"
                        print(f"Processing row {idx}: Generating {base_filename}{template_ext}...")
                        
                        template_sheet = job.metadata.get('template_sheet', None)
                        
                        self.template_processor.process_template(
                            job.metadata['job_template_path'],
                            row_data,
                            str(processed_doc),
                            sheet_name=template_sheet
                        )
                        
                        if not processed_doc.exists():
                            raise RuntimeError(f"Failed to create processed document: {processed_doc}")
                        
                        processed_docs.append({'path': processed_doc, 'priority': 0, 'template_idx': 0})
                        print(f"Row {idx}: Template processed successfully")
                    
                    # Convert to requested formats
                    for output_format in job.output_formats:
                        # Skip pdf_merged and excel_workbook formats in loop - will be handled after all files are processed
                        if output_format in ['pdf_merged', 'excel_workbook']:
                            continue
                        
                        format_dir = output_dir / output_format
                        format_dir.mkdir(parents=True, exist_ok=True)
                        
                        # Multi-template mode: convert and merge all templates
                        if len(processed_docs) > 1:
                            print(f"Row {idx}: Converting {len(processed_docs)} templates to {output_format}...")
                            
                            temp_files = []
                            for doc_info in processed_docs:
                                doc_path = doc_info['path']
                                
                                # Determine print settings
                                print_settings = None
                                if output_format == 'pdf' and job.excel_print_settings:
                                    doc_ext = Path(doc_path).suffix.lower()
                                    if doc_ext in ['.xlsx', '.xls']:
                                        print_settings = job.excel_print_settings
                                
                                try:
                                    temp_output = self.format_converter.convert(
                                        str(doc_path),
                                        output_format,
                                        str(format_dir),
                                        print_settings
                                    )
                                    
                                    if not os.path.exists(temp_output):
                                        raise RuntimeError(f"Output file was not created: {temp_output}")
                                    
                                    temp_files.append({
                                        'path': temp_output,
                                        'priority': doc_info['priority']
                                    })
                                except Exception as e:
                                    print(f"Row {idx}: Error converting template {doc_info['template_idx']}: {str(e)}")
                            
                            # Merge files if PDF format
                            if output_format == 'pdf' and len(temp_files) > 1:
                                merged_file = format_dir / f"{base_filename}.pdf"
                                self._merge_template_pdfs(temp_files, merged_file, job)
                                job.add_output_file(str(merged_file))
                                
                                # Remove temp files
                                for tf in temp_files:
                                    try:
                                        os.remove(tf['path'])
                                    except:
                                        pass
                                        
                                print(f"Row {idx}: Merged {len(temp_files)} PDFs into {merged_file.name}")
                            elif output_format == 'excel' and len(temp_files) > 1:
                                # Merge Excel files
                                merged_file = format_dir / f"{base_filename}.xlsx"
                                self._merge_template_excels(temp_files, merged_file, job)
                                job.add_output_file(str(merged_file))
                                
                                # Remove temp files
                                for tf in temp_files:
                                    try:
                                        os.remove(tf['path'])
                                    except:
                                        pass
                                        
                                print(f"Row {idx}: Merged {len(temp_files)} Excel files into {merged_file.name}")
                            else:
                                # For other formats or single file, just add to outputs
                                for tf in temp_files:
                                    job.add_output_file(tf['path'])
                        else:
                            # Single template mode (legacy)
                            processed_doc = processed_docs[0]['path']
                            print(f"Row {idx}: Converting to {output_format}...")
                            
                            print_settings = None
                            if output_format == 'pdf' and job.excel_print_settings:
                                template_ext = Path(job.metadata.get('job_template_path', '')).suffix.lower()
                                if template_ext in ['.xlsx', '.xls']:
                                    print_settings = job.excel_print_settings
                                    print(f"Row {idx}: Using Excel print settings for PDF conversion")
                            
                            try:
                                output_file = self.format_converter.convert(
                                    str(processed_doc),
                                    output_format,
                                    str(format_dir),
                                    print_settings
                                )
                                
                                if not os.path.exists(output_file):
                                    raise RuntimeError(f"Output file was not created: {output_file}")
                                
                                print(f"Row {idx}: Successfully created {output_format} file: {output_file}")
                                job.add_output_file(output_file)
                                
                            except Exception as conv_error:
                                print(f"Row {idx}: Error converting to {output_format}: {str(conv_error)}")
                                raise
                    
                    job.increment_processed()
                    print(f"Row {idx}: Completed successfully")
                    
                except Exception as e:
                    error_msg = f"Error processing row {idx}: {str(e)}"
                    print(error_msg)
                    import traceback
                    traceback.print_exc()
                    job.increment_failed()
                    if not job.error_message:
                        job.error_message = error_msg
                
                # Batched metadata saves: save every 10 rows or every 5 seconds
                rows_since_save += 1
                current_time = time.time()
                if rows_since_save >= save_count or (current_time - last_save_time) >= save_interval:
                    self.save_job_metadata(job)
                    rows_since_save = 0
                    last_save_time = current_time
            
            # Final save after loop completes
            self.save_job_metadata(job)
            
            # Validate that we have output files
            if job.processed_records == 0:
                raise RuntimeError("No records were processed successfully")
            
            # Verify output directory has files
            output_files_exist = any(output_dir.rglob('*.*'))
            if not output_files_exist:
                raise RuntimeError(f"No output files were generated in {output_dir}")
            
            # Handle PDF merging if pdf_merged format was requested
            if 'pdf_merged' in job.output_formats:
                # Check if individual PDFs were also requested
                if 'pdf' not in job.output_formats:
                    # Need to create temporary PDFs for merging
                    print(f"Job {job.id}: Creating PDFs for merging...")
                    temp_pdf_dir = output_dir / 'pdf'
                    temp_pdf_dir.mkdir(parents=True, exist_ok=True)
                    
                    # Convert all processed documents to PDF
                    # Get template extension
                    template_ext = Path(job.metadata['job_template_path']).suffix
                    
                    # Find all processed files with the template extension
                    processed_files = sorted(output_dir.glob(f"*{template_ext}"))
                    
                    if not processed_files:
                        print(f"Job {job.id}: Warning - No processed files found for PDF merging")
                    
                    for processed_file in processed_files:
                        try:
                            print_settings = None
                            if job.excel_print_settings:
                                if template_ext.lower() in ['.xlsx', '.xls']:
                                    print_settings = job.excel_print_settings
                            
                            output_file = self.format_converter.convert(
                                str(processed_file),
                                'pdf',
                                str(temp_pdf_dir),
                                print_settings
                            )
                            print(f"Job {job.id}: Created PDF for merging: {output_file}")
                        except Exception as e:
                            print(f"Job {job.id}: Error creating PDF for merging {processed_file.name}: {str(e)}")
                            import traceback
                            traceback.print_exc()
                
                print(f"Job {job.id}: Merging PDF files...")
                self._merge_pdfs(output_dir, job)
            
            # Handle Excel workbook merging if excel_workbook format was requested
            if 'excel_workbook' in job.output_formats:
                # Check if individual Excel files were also requested
                if 'excel' not in job.output_formats:
                    # Need to create temporary Excel files for merging
                    print(f"Job {job.id}: Creating Excel files for workbook merging...")
                    temp_excel_dir = output_dir / 'excel'
                    temp_excel_dir.mkdir(parents=True, exist_ok=True)
                    
                    # Get template extension
                    template_ext = Path(job.metadata['job_template_path']).suffix
                    
                    # Find all processed files with the template extension
                    processed_files = sorted(output_dir.glob(f"*{template_ext}"))
                    
                    if not processed_files:
                        print(f"Job {job.id}: Warning - No processed files found for workbook merging")
                    
                    for processed_file in processed_files:
                        try:
                            output_file = self.format_converter.convert(
                                str(processed_file),
                                'excel',
                                str(temp_excel_dir),
                                None
                            )
                            print(f"Job {job.id}: Created Excel file for workbook: {output_file}")
                        except Exception as e:
                            print(f"Job {job.id}: Error creating Excel file for workbook {processed_file.name}: {str(e)}")
                            import traceback
                            traceback.print_exc()
                
                print(f"Job {job.id}: Merging Excel files into workbook...")
                self._merge_excel_workbook(output_dir, job)
            
            print(f"Job {job.id}: Creating ZIP archive...")
            
            # Create ZIP file with all outputs
            zip_path = self.get_job_dir(job.id) / f"job_{job.id}_output.zip"
            self._create_zip_archive(output_dir, zip_path)
            
            # Verify ZIP was created
            if not zip_path.exists():
                raise RuntimeError(f"Failed to create ZIP file: {zip_path}")
            
            zip_size = zip_path.stat().st_size
            if zip_size == 0:
                raise RuntimeError(f"ZIP file is empty: {zip_path}")
            
            print(f"Job {job.id}: ZIP created successfully ({zip_size} bytes)")
            job.set_zip_file(str(zip_path))
            
            # Copy ZIP to custom output directory if specified
            if job.output_directory and os.path.exists(job.output_directory):
                try:
                    custom_zip_path = os.path.join(job.output_directory, f"job_{job.id}_output.zip")
                    shutil.copy2(zip_path, custom_zip_path)
                    print(f"Output copied to: {custom_zip_path}")
                except Exception as e:
                    print(f"Failed to copy output to custom directory: {str(e)}")
            
            # Final validation
            print(f"Job {job.id}: Processing completed. Processed: {job.processed_records}, Failed: {job.failed_records}")
            
            # Mark as completed
            job.update_status(JobStatus.COMPLETED)
            
        except Exception as e:
            job.update_status(JobStatus.FAILED, str(e))
        
        self.save_job_metadata(job)
        return job
    
    def _create_zip_archive(self, source_dir: Path, zip_path: Path):
        """Create a ZIP archive from a directory."""
        if not source_dir.exists():
            raise RuntimeError(f"Source directory does not exist: {source_dir}")
        
        file_count = 0
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(source_dir):
                for file in files:
                    file_path = Path(root) / file
                    arcname = file_path.relative_to(source_dir)
                    zipf.write(file_path, arcname)
                    file_count += 1
                    print(f"Added to ZIP: {arcname}")
        
        if file_count == 0:
            raise RuntimeError(f"No files found to archive in {source_dir}")
        
        print(f"ZIP archive created with {file_count} files")
    
    def _process_split_job(self, job: 'Job') -> 'Job':
        """
        Process a split job (PDF/Word splitting).
        
        Args:
            job: Job instance
            
        Returns:
            Updated Job instance
        """
        try:
            from services.pdf_operations import PDFSplitter
            from services.word_operations import WordSplitter
            
            # Get split configuration from metadata
            split_config = job.metadata.get('split_config', {})
            split_type = split_config.get('split_type', 'by_count')
            pages_per_split = split_config.get('pages_per_split', 1)
            names_file_path = split_config.get('names_file_path')
            input_file_path = job.metadata.get('job_data_path')  # Using data_path for input file
            
            if not input_file_path:
                raise ValueError("Input file path not found")
            
            # Create output directory
            output_dir = self.get_job_dir(job.id) / "outputs" / "splits"
            output_dir.mkdir(parents=True, exist_ok=True)
            
            # Determine file type
            file_ext = Path(input_file_path).suffix.lower()
            base_name = Path(input_file_path).stem
            
            print(f"Starting split job: {split_type}, {pages_per_split} pages per split")
            
            # Perform splitting
            output_files = []
            
            if file_ext == '.pdf':
                splitter = PDFSplitter(input_file_path)
                
                if split_type == 'by_names' and names_file_path:
                    output_files = splitter.split_by_names(names_file_path, pages_per_split, str(output_dir))
                else:
                    output_files = splitter.split_by_count(pages_per_split, str(output_dir), base_name)
            
            elif file_ext in ['.docx', '.doc']:
                splitter = WordSplitter(input_file_path)
                
                if split_type == 'by_names' and names_file_path:
                    output_files = splitter.split_by_names(names_file_path, pages_per_split, str(output_dir))
                else:
                    output_files = splitter.split_by_pages(pages_per_split, str(output_dir), base_name)
            
            else:
                raise ValueError(f"Unsupported file type for splitting: {file_ext}")
            
            # Add output files to job
            for file_path in output_files:
                job.add_output_file(file_path)
            
            job.total_records = len(output_files)
            job.processed_records = len(output_files)
            
            print(f"Split job completed: {len(output_files)} files created")
            
            # Create ZIP archive
            zip_file = self.get_job_dir(job.id) / "output.zip"
            self._create_zip_archive(output_dir, zip_file)
            job.set_zip_file(str(zip_file))
            
            job.update_status(JobStatus.COMPLETED)
            self.save_job_metadata(job)
            
            return job
        
        except Exception as e:
            error_msg = f"Error in split job: {str(e)}"
            print(error_msg)
            job.update_status(JobStatus.FAILED, error_message=error_msg)
            self.save_job_metadata(job)
            raise
    
    def _process_merge_job(self, job: 'Job') -> 'Job':
        """
        Process a merge job (PDF/Word merging).
        
        Args:
            job: Job instance
            
        Returns:
            Updated Job instance
        """
        try:
            from services.pdf_operations import PDFMerger
            from services.word_operations import WordMerger
            
            # Get merge configuration from metadata
            merge_config = job.metadata.get('merge_config', {})
            merge_mode = merge_config.get('merge_mode', 'paired')
            file_paths = merge_config.get('file_paths')
            directory_path = merge_config.get('directory_path')
            
            # Create output directory
            output_dir = self.get_job_dir(job.id) / "outputs" / "merged"
            output_dir.mkdir(parents=True, exist_ok=True)
            
            print(f"Starting merge job: {merge_mode} mode")
            
            # Determine file type and get file paths
            if merge_mode == 'paired':
                if not file_paths or len(file_paths) != 2:
                    raise ValueError("Paired merge requires exactly 2 files")
                
                file1_path = file_paths[0]
                file2_path = file_paths[1]
                
                # Determine file types
                file1_ext = Path(file1_path).suffix.lower()
                file2_ext = Path(file2_path).suffix.lower()
                
                if file1_ext != file2_ext:
                    raise ValueError(f"Cannot merge different file types: {file1_ext} and {file2_ext}")
                
                output_file = output_dir / f"merged{file1_ext}"
                
                # Perform paired merging
                if file1_ext == '.pdf':
                    merger = PDFMerger()
                    result_file = merger.merge_paired(file1_path, file2_path, str(output_file))
                elif file1_ext in ['.docx', '.doc']:
                    merger = WordMerger()
                    result_file = merger.merge_paired(file1_path, file2_path, str(output_file))
                else:
                    raise ValueError(f"Unsupported file type for merging: {file1_ext}")
            
            else:  # sequential mode
                if directory_path:
                    # Merge all files from directory
                    # Detect file type from directory contents
                    import glob
                    pdf_files = glob.glob(os.path.join(directory_path, "*.pdf"))
                    word_files = glob.glob(os.path.join(directory_path, "*.docx")) + glob.glob(os.path.join(directory_path, "*.doc"))
                    
                    if pdf_files:
                        file_ext = '.pdf'
                        output_file = output_dir / f"merged{file_ext}"
                        merger = PDFMerger()
                        result_file = merger.merge_directory(directory_path, str(output_file), file_ext)
                    elif word_files:
                        file_ext = '.docx'
                        output_file = output_dir / f"merged{file_ext}"
                        merger = WordMerger()
                        result_file = merger.merge_directory(directory_path, str(output_file))
                    else:
                        raise ValueError(f"No PDF or Word files found in directory: {directory_path}")
                
                elif file_paths and len(file_paths) > 0:
                    # Merge specific files
                    first_ext = Path(file_paths[0]).suffix.lower()
                    output_file = output_dir / f"merged{first_ext}"
                    
                    if first_ext == '.pdf':
                        merger = PDFMerger()
                        result_file = merger.merge_sequential(file_paths, str(output_file))
                    elif first_ext in ['.docx', '.doc']:
                        merger = WordMerger()
                        result_file = merger.merge_sequential(file_paths, str(output_file))
                    else:
                        raise ValueError(f"Unsupported file type for merging: {first_ext}")
                else:
                    raise ValueError("No files or directory path provided for sequential merge")
            
            # Add output file to job
            job.add_output_file(result_file)
            job.total_records = 1
            job.processed_records = 1
            
            print(f"Merge job completed: {result_file}")
            
            # Create ZIP archive
            zip_file = self.get_job_dir(job.id) / "output.zip"
            self._create_zip_archive(output_dir, zip_file)
            job.set_zip_file(str(zip_file))
            
            job.update_status(JobStatus.COMPLETED)
            self.save_job_metadata(job)
            
            return job
        
        except Exception as e:
            error_msg = f"Error in merge job: {str(e)}"
            print(error_msg)
            job.update_status(JobStatus.FAILED, error_message=error_msg)
            self.save_job_metadata(job)
            raise
    
    def get_job_output_files(self, job_id: str) -> List[str]:
        """
        Get list of output files for a job.
        
        Args:
            job_id: Job ID
            
        Returns:
            List of output file paths
        """
        job = self.get_job(job_id)
        if not job:
            return []
        return job.output_files
    
    def get_job_zip_file(self, job_id: str) -> Optional[str]:
        """
        Get ZIP file path for a job.
        
        Args:
            job_id: Job ID
            
        Returns:
            ZIP file path or None (returns absolute path)
        """
        job = self.get_job(job_id)
        if not job:
            return None
        
        if not job.zip_file_path:
            return None
        
        # Ensure the path is absolute
        zip_path = Path(job.zip_file_path)
        if not zip_path.is_absolute():
            # Convert relative path to absolute
            zip_path = self.jobs_dir.parent / job.zip_file_path
        
        return str(zip_path)
    
    def check_and_update_files(self, job_id: str) -> Dict:
        """
        Check if job files need updating and update if necessary.
        
        Args:
            job_id: Job ID
            
        Returns:
            Dictionary with update status
        """
        job = self.get_job(job_id)
        if not job:
            raise ValueError(f"Job not found: {job_id}")
        
        updates = {
            'template_updated': False,
            'data_updated': False
        }
        
        try:
            # Check template file
            if job.template_path and self.file_tracker.is_file_changed(job.template_path):
                template_info = self.file_tracker.track_file(job.template_path, force_update=True)
                job.local_template_path = template_info['local_path']
                updates['template_updated'] = True
            
            # Check data file
            if job.data_path and self.file_tracker.is_file_changed(job.data_path):
                data_info = self.file_tracker.track_file(job.data_path, force_update=True)
                job.local_data_path = data_info['local_path']
                updates['data_updated'] = True
            
            if updates['template_updated'] or updates['data_updated']:
                self.save_job_metadata(job)
        
        except Exception as e:
            updates['error'] = str(e)
        
        return updates
    
    def _merge_pdfs(self, output_dir: Path, job: Job):
        """
        Merge all PDF files from the pdf directory into a single merged.pdf.
        
        Args:
            output_dir: Output directory containing format subdirectories
            job: Job instance
        """
        from PyPDF2 import PdfMerger
        
        # Check if we have individual PDFs to merge
        pdf_dir = output_dir / 'pdf'
        if not pdf_dir.exists():
            print(f"Job {job.id}: No PDF directory found to merge")
            return
        
        # Get all PDF files sorted by name
        pdf_files = sorted(pdf_dir.glob('*.pdf'))
        if not pdf_files:
            print(f"Job {job.id}: No PDF files found to merge")
            return
        
        # Create pdf_merged directory
        merged_dir = output_dir / 'pdf_merged'
        merged_dir.mkdir(parents=True, exist_ok=True)
        merged_file = merged_dir / 'merged.pdf'
        
        try:
            # Create merger and add all PDFs
            merger = PdfMerger()
            
            for pdf_file in pdf_files:
                print(f"Job {job.id}: Adding {pdf_file.name} to merged PDF")
                merger.append(str(pdf_file))
            
            # Write merged PDF
            merger.write(str(merged_file))
            merger.close()
            
            # Verify merged file was created
            if not merged_file.exists():
                raise RuntimeError(f"Failed to create merged PDF: {merged_file}")
            
            merged_size = merged_file.stat().st_size
            print(f"Job {job.id}: Merged PDF created successfully ({merged_size} bytes, {len(pdf_files)} files merged)")
            
            # Add to job output files
            job.add_output_file(str(merged_file))
            
        except Exception as e:
            print(f"Job {job.id}: Error merging PDFs: {str(e)}")
            import traceback
            traceback.print_exc()
            # Don't fail the entire job if merge fails
    
    def _merge_template_pdfs(self, pdf_list: List[Dict], output_path: Path, job: Job):
        """
        Merge PDFs from multiple templates in priority order.
        
        Args:
            pdf_list: List of dicts with 'path' and 'priority' keys
            output_path: Path for merged output file
            job: Job instance
        """
        from PyPDF2 import PdfMerger
        
        # Sort by priority
        sorted_pdfs = sorted(pdf_list, key=lambda x: x['priority'])
        
        try:
            merger = PdfMerger()
            
            for pdf_info in sorted_pdfs:
                pdf_path = pdf_info['path']
                if os.path.exists(pdf_path):
                    merger.append(str(pdf_path))
            
            merger.write(str(output_path))
            merger.close()
            
            print(f"Job {job.id}: Merged {len(sorted_pdfs)} template PDFs into {output_path.name}")
            
        except Exception as e:
            print(f"Job {job.id}: Error merging template PDFs: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def _merge_template_excels(self, excel_list: List[Dict], output_path: Path, job: Job):
        """
        Merge Excel files from multiple templates as separate sheets.
        
        Args:
            excel_list: List of dicts with 'path' and 'priority' keys
            output_path: Path for merged output file
            job: Job instance
        """
        from openpyxl import Workbook, load_workbook
        
        # Sort by priority
        sorted_excels = sorted(excel_list, key=lambda x: x['priority'])
        
        try:
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            for idx, excel_info in enumerate(sorted_excels):
                excel_path = excel_info['path']
                if not os.path.exists(excel_path):
                    continue
                
                source_wb = load_workbook(excel_path)
                source_sheet = source_wb.active
                
                # Create new sheet with template number
                target_sheet = wb.create_sheet(title=f"Template_{idx+1}")
                
                # Copy cells
                for row in source_sheet.iter_rows():
                    for cell in row:
                        target_cell = target_sheet.cell(
                            row=cell.row,
                            column=cell.column,
                            value=cell.value
                        )
                        
                        if cell.has_style:
                            target_cell.font = cell.font.copy()
                            target_cell.border = cell.border.copy()
                            target_cell.fill = cell.fill.copy()
                            target_cell.number_format = cell.number_format
                            target_cell.protection = cell.protection.copy()
                            target_cell.alignment = cell.alignment.copy()
                
                # Copy column dimensions
                for col in source_sheet.column_dimensions:
                    target_sheet.column_dimensions[col].width = source_sheet.column_dimensions[col].width
                
                # Copy row dimensions
                for row in source_sheet.row_dimensions:
                    target_sheet.row_dimensions[row].height = source_sheet.row_dimensions[row].height
                
                source_wb.close()
            
            wb.save(str(output_path))
            wb.close()
            
            print(f"Job {job.id}: Merged {len(sorted_excels)} template Excel files into {output_path.name}")
            
        except Exception as e:
            print(f"Job {job.id}: Error merging template Excel files: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def _merge_excel_workbook(self, output_dir: Path, job: Job):
        """
        Merge all Excel files from the excel directory into a single workbook with tabs.
        
        Args:
            output_dir: Output directory containing format subdirectories
            job: Job instance
        """
        from openpyxl import Workbook, load_workbook
        
        # Check if we have individual Excel files to merge
        excel_dir = output_dir / 'excel'
        if not excel_dir.exists():
            print(f"Job {job.id}: No Excel directory found to merge")
            return
        
        # Get all Excel files sorted by name
        excel_files = sorted(excel_dir.glob('*.xlsx'))
        if not excel_files:
            print(f"Job {job.id}: No Excel files found to merge")
            return
        
        # Create excel_workbook directory
        workbook_dir = output_dir / 'excel_workbook'
        workbook_dir.mkdir(parents=True, exist_ok=True)
        workbook_file = workbook_dir / 'workbook.xlsx'
        
        try:
            # Get tabname variable from job metadata (default: ##tabname##)
            tabname_variable = job.metadata.get('tabname_variable', '##tabname##')
            
            # Get the data records to extract tab names
            data_records = job.metadata.get('data_records', [])
            
            # Create new workbook
            wb = Workbook()
            # Remove the default sheet
            wb.remove(wb.active)
            
            used_tab_names = set()
            
            # Process each Excel file
            for idx, excel_file in enumerate(excel_files):
                print(f"Job {job.id}: Adding {excel_file.name} to workbook")
                
                # Load the source workbook
                source_wb = load_workbook(excel_file)
                
                # Get the first (and usually only) sheet
                source_sheet = source_wb.active
                
                # Determine tab name
                tab_name = self._get_tab_name(
                    data_records[idx] if idx < len(data_records) else {},
                    tabname_variable,
                    idx,
                    used_tab_names
                )
                
                # Create new sheet in target workbook
                target_sheet = wb.create_sheet(title=tab_name)
                
                # Copy all cells from source to target
                for row in source_sheet.iter_rows():
                    for cell in row:
                        target_cell = target_sheet.cell(
                            row=cell.row,
                            column=cell.column,
                            value=cell.value
                        )
                        
                        # Copy cell formatting
                        if cell.has_style:
                            target_cell.font = cell.font.copy()
                            target_cell.border = cell.border.copy()
                            target_cell.fill = cell.fill.copy()
                            target_cell.number_format = cell.number_format
                            target_cell.protection = cell.protection.copy()
                            target_cell.alignment = cell.alignment.copy()
                
                # Copy column dimensions
                for col in source_sheet.column_dimensions:
                    if col in source_sheet.column_dimensions:
                        target_sheet.column_dimensions[col].width = source_sheet.column_dimensions[col].width
                
                # Copy row dimensions
                for row in source_sheet.row_dimensions:
                    if row in source_sheet.row_dimensions:
                        target_sheet.row_dimensions[row].height = source_sheet.row_dimensions[row].height
                
                # Copy merged cells
                for merged_cell_range in source_sheet.merged_cells.ranges:
                    target_sheet.merge_cells(str(merged_cell_range))
                
                source_wb.close()
            
            # Save the workbook
            wb.save(str(workbook_file))
            wb.close()
            
            # Verify workbook file was created
            if not workbook_file.exists():
                raise RuntimeError(f"Failed to create merged workbook: {workbook_file}")
            
            workbook_size = workbook_file.stat().st_size
            print(f"Job {job.id}: Merged workbook created successfully ({workbook_size} bytes, {len(excel_files)} tabs)")
            
            # Add to job output files
            job.add_output_file(str(workbook_file))
            
        except Exception as e:
            print(f"Job {job.id}: Error merging Excel workbook: {str(e)}")
            import traceback
            traceback.print_exc()
            # Don't fail the entire job if merge fails
    
    def _get_tab_name(self, data_record: Dict, tabname_variable: str, index: int, used_names: set) -> str:
        """
        Generate a valid Excel tab name from data record.
        
        Args:
            data_record: Dictionary with data for this row
            tabname_variable: Variable name to look up (e.g., '##tabname##')
            index: Row index (0-based) for fallback naming
            used_names: Set of already used tab names to avoid duplicates
            
        Returns:
            Valid Excel tab name (max 31 chars, no invalid chars, unique)
        """
        # Try to get tab name from data record
        tab_name = None
        if tabname_variable:
            # Remove ## markers to get variable name
            var_name = tabname_variable.strip('#')
            tab_name = data_record.get(var_name, '')
        
        # Fallback to Sheet{n} if no value found
        if not tab_name:
            tab_name = f"Sheet{index + 1}"
        
        # Sanitize tab name - Excel doesn't allow: \ / ? * [ ] :
        invalid_chars = ['\\', '/', '?', '*', '[', ']', ':']
        for char in invalid_chars:
            tab_name = tab_name.replace(char, '_')
        
        # Trim to max 31 characters (Excel limit)
        tab_name = tab_name[:31]
        
        # Handle duplicates by appending number
        original_name = tab_name
        counter = 1
        while tab_name in used_names:
            suffix = f"_{counter}"
            # Make sure we don't exceed 31 chars with suffix
            max_base_len = 31 - len(suffix)
            tab_name = original_name[:max_base_len] + suffix
            counter += 1
        
        used_names.add(tab_name)
        return tab_name
    
    def get_dashboard_stats(self) -> Dict:
        """
        Get statistics for dashboard display.
        
        Returns:
            Dictionary with dashboard statistics
        """
        all_jobs = self.get_all_jobs()
        
        return {
            'total_jobs': len(all_jobs),
            'pending_jobs': len([j for j in all_jobs if j.status == JobStatus.PENDING]),
            'processing_jobs': len([j for j in all_jobs if j.status == JobStatus.PROCESSING]),
            'completed_jobs': len([j for j in all_jobs if j.status == JobStatus.COMPLETED]),
            'failed_jobs': len([j for j in all_jobs if j.status == JobStatus.FAILED]),
            'total_records_processed': sum(j.processed_records for j in all_jobs),
            'total_files_generated': sum(len(j.output_files) for j in all_jobs)
        }
